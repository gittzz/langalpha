"""Shared helpers for agent-generated Excel workbooks.

Import from build scripts in work/<task>/:

    from skills.xlsx.scripts.workbook_helpers import (
        coerce_scalar,
        write_title_block,
        write_section_header,
    )
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def coerce_scalar(value: Any) -> str | int | float | bool | None:
    """Extract a cell-safe scalar from MCP/JSON payloads.

    MCP tools often return nested dicts like {"value": 0.35, "formatted": "35%"}.
    Writing the dict directly causes viewer issues; always pass values through here.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    if isinstance(value, dict):
        for key in ("value", "amount", "raw", "number", "result", "text"):
            if key in value:
                return coerce_scalar(value[key])
        if len(value) == 1:
            return coerce_scalar(next(iter(value.values())))
        raise TypeError(
            f"Cannot coerce dict to scalar — pick a field explicitly: {value!r}"
        )
    if isinstance(value, (list, tuple)):
        if len(value) == 1:
            return coerce_scalar(value[0])
        raise TypeError(f"Cannot coerce sequence to scalar: {value!r}")
    raise TypeError(f"Unsupported cell value type: {type(value).__name__}")


def _merge_row(ws: Worksheet, row: int, start_col: int, end_col: int) -> None:
    if end_col > start_col:
        ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=end_col,
        )


def write_title_block(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str = "",
    disclaimer: str = "",
    start_row: int = 1,
    start_col: int = 1,
    merge_to_col: int | None = None,
    title_font: Font | None = None,
) -> int:
    """Write document title rows ONCE and merge across columns.

    Returns the next available row index after the block.
    """
    end_col = merge_to_col or max(ws.max_column, start_col + 7)
    bold = title_font or Font(bold=True, size=12)
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.cell(row=start_row, column=start_col, value=title).font = bold
    ws.cell(row=start_row, column=start_col).alignment = align
    _merge_row(ws, start_row, start_col, end_col)

    next_row = start_row + 1
    if subtitle:
        ws.cell(row=next_row, column=start_col, value=subtitle).alignment = align
        _merge_row(ws, next_row, start_col, end_col)
        next_row += 1
    if disclaimer:
        ws.cell(row=next_row, column=start_col, value=disclaimer).alignment = align
        _merge_row(ws, next_row, start_col, end_col)
        next_row += 1
    return next_row


def write_section_header(
    ws: Worksheet,
    *,
    text: str,
    row: int,
    start_col: int = 1,
    merge_to_col: int | None = None,
) -> None:
    """Write a section label once (e.g. 'SECTION A: OPERATING METRICS')."""
    end_col = merge_to_col or max(ws.max_column, start_col + 7)
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    _merge_row(ws, row, start_col, end_col)


def last_data_column(ws: Worksheet, header_row: int, start_col: int = 1) -> int:
    """Find the rightmost non-empty header cell on a row."""
    end_col = start_col
    for col in range(start_col, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value not in (None, ""):
            end_col = col
    return end_col


def col_letter(col: int) -> str:
    return get_column_letter(col)
