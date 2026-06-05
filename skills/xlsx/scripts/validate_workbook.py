"""
Validate agent-generated Excel workbooks for common layout and formula issues.

Usage:
    python skills/xlsx/scripts/validate_workbook.py output.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def _row_text_repeats(ws, row: int, min_cols: int = 3) -> list[dict]:
    """Detect the same non-empty string copied across many columns in one row."""
    values: list[tuple[int, str]] = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is None or val == "":
            continue
        text = str(val).strip()
        if text:
            values.append((col, text))

    if len(values) < min_cols:
        return []

    texts = [t for _, t in values]
    if len(set(texts)) == 1:
        return [
            {
                "type": "repeated_row_text",
                "row": row,
                "columns": len(values),
                "text_preview": texts[0][:80],
                "hint": (
                    "Title/header was written in a column loop. Write once in column A "
                    "and use merge_cells (see workbook_helpers.write_title_block)."
                ),
            }
        ]
    return []


def validate_workbook(path: str | Path) -> dict:
    path = Path(path)
    if not path.exists():
        return {"status": "error", "error": f"File not found: {path}"}

    issues: list[dict] = []
    formula_count = 0
    formulas_without_cache_hint = 0

    wb = load_workbook(path, data_only=False)
    try:
        for ws in wb.worksheets:
            scan_rows = min(25, ws.max_row or 0)
            for row in range(1, scan_rows + 1):
                issues.extend(_row_text_repeats(ws, row))

            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        formula_count += 1
                    elif isinstance(val, dict):
                        issues.append(
                            {
                                "type": "dict_cell_value",
                                "location": f"{ws.title}!{cell.coordinate}",
                                "hint": (
                                    "Cell contains a dict object. Use "
                                    "workbook_helpers.coerce_scalar() before writing."
                                ),
                            }
                        )
    finally:
        wb.close()

    # openpyxl stores formulas as strings; cached results appear only after recalc.
    # Heuristic: if we found formulas, remind the agent to recalc.
    if formula_count > 0:
        wb_values = load_workbook(path, data_only=True)
        try:
            empty_formula_results = 0
            wb_formulas = load_workbook(path, data_only=False)
            try:
                for ws_f, ws_v in zip(wb_formulas.worksheets, wb_values.worksheets):
                    for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
                        for cell_f, cell_v in zip(row_f, row_v):
                            if (
                                isinstance(cell_f.value, str)
                                and cell_f.value.startswith("=")
                                and cell_v.value is None
                            ):
                                empty_formula_results += 1
            finally:
                wb_formulas.close()
        finally:
            wb_values.close()

        if empty_formula_results > 0:
            formulas_without_cache_hint = empty_formula_results
            issues.append(
                {
                    "type": "uncached_formulas",
                    "count": empty_formula_results,
                    "hint": (
                        "Formula cells have no cached values. Run "
                        "python skills/xlsx/scripts/recalc.py <file> before delivery — "
                        "otherwise the web viewer shows '[object Object]' for formula cells."
                    ),
                }
            )

    status = "success" if not issues else "issues_found"
    return {
        "status": status,
        "file": str(path),
        "formula_count": formula_count,
        "uncached_formula_cells": formulas_without_cache_hint,
        "issue_count": len(issues),
        "issues": issues[:50],
    }


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python validate_workbook.py <excel_file>")
        sys.exit(1)
    result = validate_workbook(sys.argv[1])
    print(json.dumps(result, indent=2))
    if result.get("status") != "success":
        sys.exit(1)


if __name__ == "__main__":
    main()
